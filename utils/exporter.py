"""
Data export utilities — CSV, Excel (XLSX), and JSON.

All exporters accept a list of Product objects and a destination path.
Files are written to the output/ directory by default.
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from parsers.models import Product
from utils.logger import setup_logger

logger = setup_logger(__name__)

OUTPUT_DIR = Path("output")


def _ensure_output_dir() -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    return OUTPUT_DIR


def _to_dataframe(products: list[Product]) -> pd.DataFrame:
    """Convert a list of Products to a clean DataFrame."""
    rows = [p.to_dict() for p in products]
    df = pd.DataFrame(rows)
    # Fill missing numeric fields with a visible placeholder
    for col in df.columns:
        if df[col].isna().any():
            df[col] = df[col].astype(object).fillna("N/A")
    return df


def _timestamped_stem(keyword: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = keyword.replace(" ", "_").lower()[:30]
    return f"{safe}_{ts}"


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def export_csv(
    products: list[Product],
    keyword: str,
    out_dir: Optional[Path] = None,
) -> Path:
    """Write products to a UTF-8 CSV file and return the file path."""
    dest = (out_dir or _ensure_output_dir()) / f"{_timestamped_stem(keyword)}.csv"
    df = _to_dataframe(products)
    df.to_csv(dest, index=False, encoding="utf-8-sig")
    logger.info("CSV exported → %s (%d rows)", dest, len(df))
    return dest


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def export_json(
    products: list[Product],
    keyword: str,
    out_dir: Optional[Path] = None,
) -> Path:
    """Write products to a pretty-printed JSON file and return the path."""
    dest = (out_dir or _ensure_output_dir()) / f"{_timestamped_stem(keyword)}.json"
    payload = {
        "keyword": keyword,
        "scraped_at": datetime.now().isoformat(),
        "total": len(products),
        "products": [p.to_full_dict() for p in products],
    }
    dest.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("JSON exported → %s (%d items)", dest, len(products))
    return dest


# ---------------------------------------------------------------------------
# Excel (XLSX) — professionally formatted
# ---------------------------------------------------------------------------

# Colour palette
_HDR_FILL   = PatternFill("solid", start_color="1F4E79")   # dark-blue header
_ALT_FILL   = PatternFill("solid", start_color="D9E1F2")   # light-blue alt rows
_WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
_HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BODY_FONT  = Font(name="Arial", size=10)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN_SIDE  = Side(style="thin", color="B8B8B8")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)

# Column definitions: (header, min_width, alignment)
_COLUMNS: list[tuple[str, int, Alignment]] = [
    ("#",             5,  _CENTER),
    ("Source",       10,  _CENTER),
    ("Title",        45,  _LEFT),
    ("Price",        12,  _CENTER),
    ("Currency",      9,  _CENTER),
    ("Rating",        9,  _CENTER),
    ("Reviews",      10,  _CENTER),
    ("Availability", 14,  _CENTER),
    ("URL",          40,  _LEFT),
]


def export_excel(
    products: list[Product],
    keyword: str,
    out_dir: Optional[Path] = None,
) -> Path:
    """
    Write products to a formatted XLSX workbook and return the file path.

    Sheet layout
    ~~~~~~~~~~~~
    Row 1  — Title banner
    Row 2  — Metadata (keyword, timestamp, total count)
    Row 3  — Blank spacer
    Row 4  — Column headers
    Row 5+ — Data rows (alternating fill)
    Last   — Summary row (count, avg price, avg rating)
    """
    dest = (out_dir or _ensure_output_dir()) / f"{_timestamped_stem(keyword)}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    num_cols = len(_COLUMNS)
    last_col_letter = get_column_letter(num_cols)

    # ── Row 1: Title banner ────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"🛍  E-commerce Tracker — Search: \"{keyword}\""
    title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="0D3B66")
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 30

    # ── Row 2: Metadata ────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col_letter}2")
    meta = ws["A2"]
    meta.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   "
        f"Total products: {len(products)}"
    )
    meta.font = Font(name="Arial", italic=True, size=10, color="555555")
    meta.alignment = _CENTER
    meta.fill = PatternFill("solid", start_color="EBF0F7")
    ws.row_dimensions[2].height = 18

    # ── Row 3: Spacer ──────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Headers ─────────────────────────────────────────────────
    for col_idx, (header, width, _align) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 22

    # ── Rows 5+: Data ──────────────────────────────────────────────────
    data_start = 5
    for row_idx, product in enumerate(products, start=data_start):
        fill = _ALT_FILL if (row_idx - data_start) % 2 == 0 else _WHITE_FILL
        row_values = [
            row_idx - data_start + 1,
            product.source.capitalize(),
            product.title,
            product.price if product.price is not None else "N/A",
            product.currency,
            product.rating if product.rating is not None else "N/A",
            product.review_count if product.review_count is not None else "N/A",
            product.availability,
            product.url,
        ]
        aligns = [a for _, _, a in _COLUMNS]
        for col_idx, (value, align) in enumerate(zip(row_values, aligns), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.fill = fill
            cell.alignment = align
            cell.border = _THIN_BORDER
        ws.row_dimensions[row_idx].height = 18

    # ── Summary row ────────────────────────────────────────────────────
    summary_row = data_start + len(products)
    ws.row_dimensions[summary_row].height = 20

    # Count
    count_cell = ws.cell(row=summary_row, column=1, value=f"Total: {len(products)}")
    count_cell.font = Font(name="Arial", bold=True, size=10)
    count_cell.fill = PatternFill("solid", start_color="1F4E79")
    count_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    count_cell.alignment = _CENTER
    count_cell.border = _THIN_BORDER

    # Avg price formula (column D = col 4)
    price_col = get_column_letter(4)
    avg_price_cell = ws.cell(
        row=summary_row, column=4,
        value=f"=IFERROR(AVERAGEIF({price_col}{data_start}:{price_col}{summary_row-1},\"<>N/A\"),\"N/A\")",
    )
    avg_price_cell.number_format = "#,##0.00"
    avg_price_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    avg_price_cell.fill = PatternFill("solid", start_color="1F4E79")
    avg_price_cell.alignment = _CENTER
    avg_price_cell.border = _THIN_BORDER

    # Avg rating formula (column F = col 6)
    rating_col = get_column_letter(6)
    avg_rating_cell = ws.cell(
        row=summary_row, column=6,
        value=f"=IFERROR(AVERAGEIF({rating_col}{data_start}:{rating_col}{summary_row-1},\"<>N/A\"),\"N/A\")",
    )
    avg_rating_cell.number_format = "0.0"
    avg_rating_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    avg_rating_cell.fill = PatternFill("solid", start_color="1F4E79")
    avg_rating_cell.alignment = _CENTER
    avg_rating_cell.border = _THIN_BORDER

    # Fill remaining summary cells
    for col_idx in range(2, num_cols + 1):
        if col_idx in (4, 6):
            continue
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.border = _THIN_BORDER

    # Freeze header rows
    ws.freeze_panes = "A5"

    wb.save(dest)
    logger.info("Excel exported → %s (%d rows)", dest, len(products))
    return dest
